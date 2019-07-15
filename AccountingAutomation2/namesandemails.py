
import requests
import datetime
import openpyxl
from openpyxl import Workbook
from requests.auth import HTTPBasicAuth

key = '64b48167aca25c0c337e493defe30678'
secret = '3f72cc66b271b750886ae34465352cc9'
url = 'https://api.guesty.com/api/v2/guests'
params = {
    'content-type' : 'application/json', 
    'limit' : '100', 
  }
NICK = 'noName'
GuestLibrary = []
skips = True
skipAmount = 0
while skips != False:
    params['skip'] = str(skipAmount)
    connection = requests.get(url, params=params, auth=HTTPBasicAuth(key, secret))
    response = connection.json()
    if response['results'] == []:
        skips = False
    else:
        skipAmount = skipAmount + 100
    GuestLibrary = GuestLibrary + response['results']

Emailsp = []
for k in range(0,len(GuestLibrary)):
	Emailsp.append([])
	em = GuestLibrary[k]['emails']
	for j in em:
		if (not 'airbnb' in j) and (not 'homeaway' in j) and (not 'booking.com' in j):
			Emailsp[k].append(j)

GuestEm = []
Guests = []
for k in range(0,len(GuestLibrary)):
    try:
            H = ''
            for j in Emailsp[k]:
                H = H + j + '  '
            GuestEm.append(H)
    except:
        GuestEm.append('Fail@Fail.fail')
    try:
        Guests.append([GuestEm[k],GuestLibrary[k]['firstName'],GuestLibrary[k]['lastName']])
    except:
        Guests.append([GuestEm[k],'None','None'])

wb = Workbook()
ws = wb.active
ws['A1'] = 'Email'
ws['B1'] = 'First Name'
ws['C1'] = 'Last Name'
for k in range(0,len(GuestEm)):
    ws['A'+str(k+2)] = Guests[k][0]
    ws['B'+str(k+2)] = Guests[k][1]
    ws['C'+str(k+2)] = Guests[k][2]

for cell in ws['A']:
    if cell.value == '':
        ws.delete_rows(cell.row)

wb.save('GuestList.xlsx')
